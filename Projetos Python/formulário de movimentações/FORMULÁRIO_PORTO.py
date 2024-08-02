from tkinter import *
from openpyxl import  Workbook, load_workbook
from datetime import datetime
import locale
import os


#JANELA
janela = Tk()
janela.title("Formulário de movimentação")
janela.geometry("1080x200")
janela.resizable(width=False, height=False)

#LISTAS

cliente = StringVar()
container = StringVar()
tipo = StringVar()
lacre = StringVar()
reserva = StringVar()
motorista = StringVar()


###############################INTERFACE###################################################################
###########CLIENTE_CIMA#############
cliente_cima = Label(janela,text= 'Cliente')
cliente_cima.grid(column=0 , row= 1, pady = 20, padx= 20)
cliente_cx = Entry(janela, bd =2, width = 20, textvariable=cliente , justify='center')
cliente_cx.grid(column=0,row=2,padx=20)
###########CONTAINER################
container_cima = Label(janela,text='Container')
container_cima.grid(column=1,row=1,padx=50)
container_cx = Entry(janela, bd =2, width = 20, textvariable=container, justify='center')
container_cx.grid(column=1,row=2,padx=50)
###########TIPO#####################
tipo_cima = Label(janela, text = 'Tipo ')
tipo_cima.grid(column=2,row=1,padx=50)
tipo_cx = Entry(janela, bd =2, width = 20, textvariable=tipo, justify='center')
tipo_cx.grid(column=2, row = 2)
###########LACRE####################
lacre_cima = Label(janela,text ='Lacre')
lacre_cima.grid(column=3, row = 1, padx = 50)
lacre_cx = Entry(janela, bd =2, width = 20, textvariable=lacre, justify='center')
lacre_cx.grid(column=3,row = 2, padx=50 )
###########RESERVA##################
reserva_cima = Label(janela, text= 'Reserva')
reserva_cima.grid(column=4 , row= 1)
reserva_cx = Entry(janela, bd =2, width = 20, textvariable=reserva, justify='center')
reserva_cx.grid(column=4, row = 2 )
###########MOTORISTA################
motorista_cima = Label(janela, text ='Motorista')
motorista_cima.grid(column=5 , row = 1, padx = 50)
motorista_cx = Entry(janela, bd =2, width = 20, textvariable=motorista, justify='center')
motorista_cx.grid(column=5 , row = 2, padx = 50)
###############################################################################################################
#Carrega o arquivo existente

f_movimentacao = ''

locale.setlocale(locale.LC_ALL, 'pt_BR.utf8')
dir_path = os.path.dirname(os.path.realpath(__file__))

data_hoje = datetime.now()
data_hoje = data_hoje.strftime("%d-%B-%Y")

f_movimentacao = f'{dir_path}\\projetos_python{data_hoje}-f_movimentacao.xlsx'

if not os.path.isdir (f'{dir_path}\\projetos_python'):
    diretorio = f'{dir_path}\\projetos_python'
    os.makedirs(diretorio)

################################DEF############################################################################

def salvar1(cliente, container, tipo, lacre, reserva, motorista):

    locale.setlocale(locale.LC_ALL, 'pt_BR.utf8')
    dir_path = os.path.dirname(os.path.realpath(__file__))

    data_hoje = datetime.now()
    data_hoje = data_hoje.strftime("%d-%B-%Y")

    nome_arquivo = f'{dir_path}\\planilhas\\{data_hoje}-movimentações.xlsx'

    wb = ''
    if not os.path.isdir (f'{dir_path}\\planilhas'):
        diretorio = f'{dir_path}\\planilhas'
        os.makedirs(diretorio)
    try:
    
        wb = load_workbook(nome_arquivo)
        planilha = wb.active
    except:
       
        wb = Workbook()
        planilha = wb.active
        cabecalho = ["Cliente","Container", "Tipo", "Lacre", "Reserva","Motorista"]
        planilha.append(cabecalho)
      
   
    linha  = [cliente, container, tipo, lacre, reserva, motorista]
   
    
    planilha.append(linha)
       
    wb.save(nome_arquivo)

    
def salvar2():
    salvar1(cliente.get(), container.get(), tipo.get(), lacre.get(), reserva.get(), motorista.get())

###################BOTÕES##########################################################

botao_salvar = Button(janela, text='Salvar', command=salvar2)
botao_salvar.grid(column= 2, row = 10, pady= 50)
botao_limpar = Button(janela, text = 'Limpar', command = '' )
botao_limpar.grid(column = 3, row = 10 , pady = 50)












# inicializa a janela
janela.mainloop()